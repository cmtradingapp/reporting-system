"""
QA report generation: Excel (openpyxl) + PDF (reportlab).
"""
import os
from datetime import datetime
from typing import List
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT

from qa.checks.base import QAResult, STATUS

# ── colour palettes ──────────────────────────────────────────────────────────
_FILL = {
    STATUS["PASS"]:  PatternFill("solid", fgColor="C6EFCE"),  # green
    STATUS["WARN"]:  PatternFill("solid", fgColor="FFEB9C"),  # yellow
    STATUS["FAIL"]:  PatternFill("solid", fgColor="FFC7CE"),  # red
    STATUS["ERROR"]: PatternFill("solid", fgColor="E4BFFF"),  # purple
}
_STATUS_EMOJI = {
    STATUS["PASS"]: "✅", STATUS["WARN"]: "⚠️",
    STATUS["FAIL"]: "❌", STATUS["ERROR"]: "🔴",
}

_THIN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def _header_row(ws, cols, row=1):
    for c, val in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2F4F8F")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = _THIN


def _auto_width(ws, min_w=8, max_w=50):
    for col in ws.columns:
        w = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(w + 2, min_w), max_w)


# ── Excel ────────────────────────────────────────────────────────────────────

def write_excel(results: List[QAResult], score: float, folder: str, date_str: str) -> str:
    wb = Workbook()
    wb.remove(wb.active)

    _sheet_summary(wb, results, score, date_str)
    _sheet_all_checks(wb, results)
    _sheet_anomalies(wb, results)
    _sheet_score_trend(wb, folder, date_str, score, results)

    path = os.path.join(folder, f"qa_report_{date_str}.xlsx")
    wb.save(path)
    return path


def _sheet_summary(wb, results: List[QAResult], score: float, date_str: str):
    ws = wb.create_sheet("Summary")
    ws.freeze_panes = "A3"

    # Title
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = f"QA Report — {date_str}   |   Overall Score: {round(score,1)}%"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")

    # Group by report + section
    groups = defaultdict(list)
    for r in results:
        groups[(r.report, r.section)].append(r)

    cols = ["Report", "Section", "Total", "PASS", "WARN", "FAIL", "ERROR", "Score%", "Status"]
    _header_row(ws, cols, row=2)

    for row_idx, ((rep, sec), items) in enumerate(sorted(groups.items()), start=3):
        total  = len(items)
        passes = sum(1 for x in items if x.status == STATUS["PASS"])
        warns  = sum(1 for x in items if x.status == STATUS["WARN"])
        fails  = sum(1 for x in items if x.status == STATUS["FAIL"])
        errors = sum(1 for x in items if x.status == STATUS["ERROR"])
        s      = round(passes / total * 100, 1) if total else 0
        status = (STATUS["ERROR"] if errors > 0 else
                  STATUS["FAIL"]  if fails  > 0 else
                  STATUS["WARN"]  if warns  > 0 else STATUS["PASS"])
        row = [rep, sec, total, passes, warns, fails, errors, f"{s}%", _STATUS_EMOJI[status]]
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=c, value=val)
            cell.fill = _FILL[status]
            cell.border = _THIN

    _auto_width(ws)


def _sheet_all_checks(wb, results: List[QAResult]):
    ws = wb.create_sheet("All Checks")
    ws.freeze_panes = "A2"
    cols = ["Report", "Section", "Check", "Context", "Expected", "Actual", "Diff", "Pct Diff", "Status", "Message"]
    _header_row(ws, cols)
    for i, r in enumerate(results, start=2):
        row = [
            r.report, r.section, r.check_name, r.context,
            r.expected, r.actual, r.diff, f"{round(r.pct_diff*100,3)}%",
            r.status, r.message,
        ]
        fill = _FILL[r.status]
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=i, column=c, value=val)
            cell.fill = fill
            cell.border = _THIN
    _auto_width(ws)


def _sheet_anomalies(wb, results: List[QAResult]):
    ws = wb.create_sheet("Anomalies")
    ws.freeze_panes = "A2"
    anomalies = [r for r in results if r.status in (STATUS["FAIL"], STATUS["ERROR"])]
    if not anomalies:
        ws["A1"] = "No anomalies found ✅"
        return
    cols = ["Report", "Section", "Check", "Context", "Expected", "Actual", "Diff", "Status", "Message"]
    _header_row(ws, cols)
    for i, r in enumerate(anomalies, start=2):
        row = [r.report, r.section, r.check_name, r.context, r.expected, r.actual, r.diff, r.status, r.message]
        fill = _FILL[r.status]
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=i, column=c, value=val)
            cell.fill = fill
            cell.border = _THIN
    _auto_width(ws)


def _sheet_score_trend(wb, folder: str, date_str: str, score: float, results: List[QAResult]):
    from qa.history import load_history
    ws = wb.create_sheet("Score Trend")
    ws.freeze_panes = "A2"
    cols = ["Date", "Score%", "Total", "Pass", "Warn", "Fail", "Error"]
    _header_row(ws, cols)

    history = load_history(folder)
    # Include today's run
    passes  = sum(1 for r in results if r.status == STATUS["PASS"])
    fails   = sum(1 for r in results if r.status == STATUS["FAIL"])
    warns   = sum(1 for r in results if r.status == STATUS["WARN"])
    errors  = sum(1 for r in results if r.status == STATUS["ERROR"])
    history[date_str] = {
        "score": round(score,2), "total": len(results),
        "pass": passes, "warn": warns, "fail": fails, "error": errors,
    }

    data_rows = []
    for date_key in sorted(history.keys()):
        h = history[date_key]
        data_rows.append([
            date_key, h.get("score",0), h.get("total",0),
            h.get("pass",0), h.get("warn",0), h.get("fail",0), h.get("error",0),
        ])

    for i, row in enumerate(data_rows, start=2):
        for c, val in enumerate(row, 1):
            ws.cell(row=i, column=c, value=val).border = _THIN

    # Bar chart for score trend
    if len(data_rows) >= 2:
        chart = BarChart()
        chart.title = "QA Score Trend"
        chart.y_axis.title = "Score %"
        chart.x_axis.title = "Date"
        chart.style = 10
        chart.width = 20
        chart.height = 10
        data_ref   = Reference(ws, min_col=2, min_row=1, max_row=1+len(data_rows))
        cats_ref   = Reference(ws, min_col=1, min_row=2, max_row=1+len(data_rows))
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        ws.add_chart(chart, f"A{len(data_rows)+4}")

    _auto_width(ws)


# ── PDF ──────────────────────────────────────────────────────────────────────

def write_pdf(results: List[QAResult], score: float, folder: str, date_str: str) -> str:
    path = os.path.join(folder, f"qa_report_{date_str}.pdf")
    doc  = SimpleDocTemplate(path, pagesize=landscape(A4),
                              leftMargin=1.5*cm, rightMargin=1.5*cm,
                              topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    normal = styles["Normal"]
    h1     = styles["Heading1"]
    h2     = styles["Heading2"]

    elements = []

    # ── Page 1: Title + Summary ──────────────────────────────────────────────
    run_time = datetime.now().strftime("%Y-%m-%d %H:%M")
    elements.append(Paragraph(f"Data QA Validation Report — {date_str}", h1))
    elements.append(Paragraph(f"Generated: {run_time}", normal))
    elements.append(Spacer(1, 0.3*cm))

    score_color = "#27AE60" if score >= 90 else "#E67E22" if score >= 75 else "#E74C3C"
    badge_style = ParagraphStyle("badge", parent=normal, fontSize=18,
                                  textColor=score_color, alignment=TA_CENTER)
    elements.append(Paragraph(f"Overall Score: {round(score,1)}%", badge_style))
    elements.append(Spacer(1, 0.5*cm))

    # Summary table
    groups = defaultdict(list)
    for r in results:
        groups[(r.report, r.section)].append(r)

    summary_data = [["Report", "Section", "Total", "Pass", "Warn", "Fail", "Error", "Score%"]]
    for (rep, sec), items in sorted(groups.items()):
        total  = len(items)
        passes = sum(1 for x in items if x.status == STATUS["PASS"])
        warns  = sum(1 for x in items if x.status == STATUS["WARN"])
        fails  = sum(1 for x in items if x.status == STATUS["FAIL"])
        errors = sum(1 for x in items if x.status == STATUS["ERROR"])
        s      = round(passes / total * 100, 1) if total else 0
        summary_data.append([rep, sec, total, passes, warns, fails, errors, f"{s}%"])

    t = Table(summary_data, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#2F4F8F")),
        ("TEXTCOLOR",  (0,0), (-1,0), colors.white),
        ("FONTNAME",   (0,0), (-1,0), "Helvetica-Bold"),
        ("GRID",       (0,0), (-1,-1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#F0F4FF")]),
        ("ALIGN",      (2,0), (-1,-1), "CENTER"),
        ("FONTSIZE",   (0,0), (-1,-1), 9),
    ]))
    elements.append(t)
    elements.append(PageBreak())

    # ── Pages 2+: Per-report sections ────────────────────────────────────────
    for report_name in ["Performance", "Agent Bonuses", "Dashboard", "FTC Date", "Sync"]:
        report_items = [r for r in results if r.report == report_name]
        if not report_items:
            continue
        elements.append(Paragraph(report_name, h2))
        elements.append(Spacer(1, 0.2*cm))

        tbl_data = [["Section", "Check", "Context", "Expected", "Actual", "Diff", "Status", "Message"]]
        for r in report_items:
            msg_short = (r.message[:60] + "…") if len(r.message) > 60 else r.message
            tbl_data.append([
                r.section, r.check_name, r.context,
                str(r.expected)[:20], str(r.actual)[:20],
                round(r.diff, 3), r.status, msg_short,
            ])

        t = Table(tbl_data, repeatRows=1,
                  colWidths=[2.5*cm, 3*cm, 3*cm, 2.5*cm, 2.5*cm, 1.8*cm, 1.5*cm, None])
        style = [
            ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#2F4F8F")),
            ("TEXTCOLOR",  (0,0), (-1,0), colors.white),
            ("FONTNAME",   (0,0), (-1,0), "Helvetica-Bold"),
            ("GRID",       (0,0), (-1,-1), 0.4, colors.grey),
            ("FONTSIZE",   (0,0), (-1,-1), 7.5),
            ("VALIGN",     (0,0), (-1,-1), "TOP"),
        ]
        for row_i, r in enumerate(report_items, start=1):
            bg = {
                STATUS["PASS"]:  colors.HexColor("#E8F5E9"),
                STATUS["WARN"]:  colors.HexColor("#FFF9C4"),
                STATUS["FAIL"]:  colors.HexColor("#FFCDD2"),
                STATUS["ERROR"]: colors.HexColor("#EDE7F6"),
            }.get(r.status, colors.white)
            style.append(("BACKGROUND", (0, row_i), (-1, row_i), bg))
        t.setStyle(TableStyle(style))
        elements.append(t)
        elements.append(Spacer(1, 0.4*cm))

    # ── Last page: Anomalies ─────────────────────────────────────────────────
    elements.append(PageBreak())
    elements.append(Paragraph("Anomalies (FAIL + ERROR)", h2))
    anomalies = [r for r in results if r.status in (STATUS["FAIL"], STATUS["ERROR"])]
    if not anomalies:
        elements.append(Paragraph("No anomalies found.", normal))
    else:
        anom_data = [["Report", "Section", "Check", "Context", "Expected", "Actual", "Status", "Message"]]
        for r in anomalies:
            msg_short = (r.message[:70] + "…") if len(r.message) > 70 else r.message
            anom_data.append([
                r.report, r.section, r.check_name, r.context,
                str(r.expected)[:20], str(r.actual)[:20], r.status, msg_short,
            ])
        t = Table(anom_data, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#8B0000")),
            ("TEXTCOLOR",  (0,0), (-1,0), colors.white),
            ("FONTNAME",   (0,0), (-1,0), "Helvetica-Bold"),
            ("GRID",       (0,0), (-1,-1), 0.4, colors.grey),
            ("FONTSIZE",   (0,0), (-1,-1), 8),
            ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.HexColor("#FFCDD2"), colors.HexColor("#FFEBEE")]),
        ]))
        elements.append(t)

    # Score trend table on last page
    elements.append(Spacer(1, 0.6*cm))
    elements.append(Paragraph("Score Trend", h2))
    from qa.history import load_history
    history = load_history(folder)
    if history:
        trend_data = [["Date", "Score%", "Total", "Pass", "Warn", "Fail", "Error"]]
        for date_key in sorted(history.keys()):
            h = history[date_key]
            trend_data.append([
                date_key, f"{h.get('score',0)}%", h.get("total",0),
                h.get("pass",0), h.get("warn",0), h.get("fail",0), h.get("error",0),
            ])
        t = Table(trend_data, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#2F4F8F")),
            ("TEXTCOLOR",  (0,0), (-1,0), colors.white),
            ("FONTNAME",   (0,0), (-1,0), "Helvetica-Bold"),
            ("GRID",       (0,0), (-1,-1), 0.4, colors.grey),
            ("FONTSIZE",   (0,0), (-1,-1), 8),
            ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#F0F4FF")]),
        ]))
        elements.append(t)

    doc.build(elements)
    return path
